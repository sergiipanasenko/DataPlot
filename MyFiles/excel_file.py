from os.path import getmtime
from openpyxl import load_workbook
from xlrd import open_workbook as open_old_workbook
from pyxlsb import open_workbook as open_binary_workbook
from .my_file import MyFile


class MyExcelFile(MyFile):
    def __init__(self, file_name=None):
        MyFile.__init__(self, file_name)
        self.data = dict()

    def read_new_book(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
            wb = load_workbook(self.file_name)
            for sheet in wb:
                self.data[sheet.title] = []
                for row in range(1, sheet.max_row + 1):
                    current_row = []
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row, col).value
                        if cell_value is not None:
                            current_row.append(str(cell_value))
                        else:
                            current_row.append('')
                    self.data[sheet.title].append(current_row)
            wb.close()

    def read_old_book(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
            with open_old_workbook(self.file_name) as wb:
                for sheet in wb:
                    self.data[sheet.name] = []
                    for row in range(sheet.nrows):
                        current_row = []
                        for col in range(sheet.ncols):
                            cell_value = sheet.cell_value(row, col)
                            if cell_value is not None:
                                current_row.append(str(cell_value))
                            else:
                                current_row.append('')
                        self.data[sheet.name].append(current_row)

    def read_binary_book(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
            with open_binary_workbook(self.file_name) as wb:
                for sheet_name in wb.sheets:
                    with wb.get_sheet(sheet_name) as sheet:
                        self.data[sheet_name] = []
                        for row in sheet.rows():
                            current_row = []
                            for cell in row:
                                if cell.v is not None:
                                    current_row.append(str(cell.v))
                                else:
                                    current_row.append('')
                            self.data[sheet_name].append(current_row)