from openpyxl import load_workbook
from xlrd import open_workbook as open_old_workbook
from pyxlsb import open_workbook as open_binary_workbook
from .base_file import MyFile, IData, IFileType


class MyNewExcelFile(MyFile, IData):
    def __init__(self, file_name=None):
        super().__init__(file_name)

    def read_data(self, file_name=None):
        file_name = self.check_file_name(file_name)
        wb = load_workbook(file_name)
        data = dict()
        for sheet in wb:
            data[sheet.title] = []
            for row in range(1, sheet.max_row + 1):
                current_row = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row, col).value
                    if cell_value is not None:
                        current_row.append(str(cell_value))
                    else:
                        current_row.append('')
                data[sheet.title].append(current_row)
        wb.close()
        self.set_data(data)


class MyOldExcelFile(MyFile, IData):
    def __init__(self, file_name=None):
        super().__init__(file_name)

    def read_data(self, file_name=None):
        file_name = self.check_file_name(file_name)
        data = dict()
        with open_old_workbook(file_name) as wb:
            for sheet in wb:
                data[sheet.name] = []
                for row in range(sheet.nrows):
                    current_row = []
                    for col in range(sheet.ncols):
                        cell_value = sheet.cell_value(row, col)
                        if cell_value is not None:
                            current_row.append(str(cell_value))
                        else:
                            current_row.append('')
                    data[sheet.name].append(current_row)
        self.set_data(data)


class MyBinaryExcelFile(MyFile, IData):
    def __init__(self, file_name=None):
        super().__init__(file_name)

    def read_data(self, file_name=None):
        file_name = self.check_file_name(file_name)
        data = dict()
        with open_binary_workbook(file_name) as wb:
            for sheet_name in wb.sheets:
                with wb.get_sheet(sheet_name) as sheet:
                    data[sheet_name] = []
                    for row in sheet.rows():
                        current_row = []
                        for cell in row:
                            if cell.v is not None:
                                current_row.append(str(cell.v))
                            else:
                                current_row.append('')
                        data[sheet_name].append(current_row)
        self.set_data(data)


excel_file_matching = {
    'xlsx': MyNewExcelFile,
    'xlsm': MyNewExcelFile,
    'xltx': MyNewExcelFile,
    'xltm': MyNewExcelFile,
    'xls':  MyOldExcelFile,
    'xlsb': MyBinaryExcelFile
}


class MyExcelFile(MyFile, IData, IFileType):
    def __init__(self, file_name=None, excel_file_type=None):
        super().__init__(file_name)
        self.set_type_dict(excel_file_matching)
        if excel_file_type is not None:
            self.check_file_type(excel_file_type)
        self.set_file_type(excel_file_type)

    def read_data(self, file_name=None, excel_file_type=None):
        file_name = self.check_file_name(file_name)
        if excel_file_type is None:
            excel_file_type = file_name.split('.')[-1]
        else:
            self.check_file_type(excel_file_type)
        excel_file = self.get_type_dict()[excel_file_type](file_name)
        excel_file.read_data()
        self.set_data(excel_file.get_data())
        self.set_file_type(excel_file_type)
