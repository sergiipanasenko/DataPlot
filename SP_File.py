from os.path import exists, getmtime
from openpyxl import load_workbook
from xlrd import open_workbook


class MyFile:
    def __init__(self, file_name=None):
        super().__init__()
        self.file_name = file_name
        self.mod_time = None

    def _change_file_name(self, file_name):
        if file_name:
            self.file_name = file_name

    def is_exist(self, file_name=None):
        self._change_file_name(file_name)
        if exists(self.file_name):
            return True
        else:
            return False


class MyTextFile(MyFile):
    def __init__(self, file_name=None):
        MyFile.__init__(self, file_name)
        self.text = ''

    def read_text(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
            with open(self.file_name) as file:
                self.text = file.read()
                self.mod_time = getmtime(self.file_name)

    def write_text(self, text: str, marker='a', file_name=None):
        self._change_file_name(file_name)
        if marker != 'a' and marker != 'w':
            marker = 'a'
        with open(self.file_name, marker) as file:
            file.write(text)


class MyDataFile(MyTextFile):
    def __init__(self, file_name=None):
        MyTextFile.__init__(self, file_name)
        self.data = None

    def read_data(self, file_name=None):
        self.read_text(file_name)
        self.text = self.text.strip()
        self.data = []
        for row in self.text.splitlines():
            row_data = row.rstrip().split()
            self.data.append(row_data)

    def write_data(self, data: list, marker='a', file_name=None):
        self._change_file_name(file_name)
        if marker != 'a' and marker != 'w':
            marker = 'a'
        with open(self.file_name, marker) as file:
            for row in data:
                for item in row:
                    file.write(item)
                file.write('\n')


class MyExcelFile(MyFile):
    def __init__(self, file_name=None):
        MyFile.__init__(self, file_name)
        self.data = dict()

    def read_new_book(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
            wb = load_workbook(self.file_name)
            for sheet in wb:
                self.data[sheet.title] = []
                for row in range(1, sheet.max_row + 1):
                    current_row = []
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row, col).value
                        if cell_value:
                            current_row.append(str(cell_value))
                        else:
                            current_row.append('')
                    self.data[sheet.title].append(current_row)

    def read_old_book(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
            wb = open_workbook(self.file_name)
            for sheet in wb:
                self.data[sheet.name] = []
                for row in range(sheet.nrows):
                    current_row = []
                    for col in range(sheet.ncols):
                        cell_value = sheet.cell_value(row, col)
                        if cell_value:
                            current_row.append(str(cell_value))
                        else:
                            current_row.append('')
                    self.data[sheet.name].append(current_row)

    def read_binary_book(self, file_name=None):
        if (self.file_name != file_name) or (self.mod_time != getmtime(file_name)):
            self._change_file_name(file_name)
